"""
Reports Views - EIAMS (Enterprise Upgrade)
==========================================
Full report views with filters, chart data, and Excel/CSV/PDF exports.
"""

import csv
import io
from datetime import date, timedelta

from django.contrib.auth.decorators import login_required
from django.db.models import Count, F, Q, Sum
from django.http import HttpResponse
from django.shortcuts import render
from django.utils import timezone

from apps.accounts.decorators import manager_or_above_required


# ──────────────────────────────────────────────────────────────────────────────
# HELPERS
# ──────────────────────────────────────────────────────────────────────────────

def _get_generated_by(request):
    return getattr(request.user, 'full_name', request.user.username)


def _parse_date(value, fallback):
    try:
        return date.fromisoformat(value)
    except (ValueError, TypeError):
        return fallback


# ──────────────────────────────────────────────────────────────────────────────
# REPORT DASHBOARD (replaces report_index)
# ──────────────────────────────────────────────────────────────────────────────

@login_required
@manager_or_above_required
def report_dashboard(request):
    """Aggregated KPI dashboard for reports."""
    from apps.assets.models import Asset
    from apps.inventory.models import InventoryItem
    from apps.stock.models import StockMovement

    today = timezone.now().date()
    thirty_days_ago = today - timedelta(days=30)

    # ── KPI cards ────────────────────────────────────────────────────────────
    total_items  = InventoryItem.objects.filter(status='Active').count()
    total_assets = Asset.objects.filter(is_active=True).count()
    total_inv_value = (
        InventoryItem.objects.filter(status='Active')
        .aggregate(v=Sum(F('purchase_price') * F('current_qty')))['v'] or 0
    )
    low_stock_count = InventoryItem.objects.filter(
        status='Active', current_qty__lte=F('min_qty')
    ).count()

    # ── Stock totals ─────────────────────────────────────────────────────────
    stock_in_types  = ['Stock IN', 'Purchase']
    stock_out_types = ['Stock OUT', 'Damage', 'Lost', 'Expired', 'Usage']

    stock_in_total  = (
        StockMovement.objects.filter(movement_type__in=stock_in_types)
        .aggregate(t=Sum('quantity'))['t'] or 0
    )
    stock_out_total = (
        StockMovement.objects.filter(movement_type__in=stock_out_types)
        .aggregate(t=Sum('quantity'))['t'] or 0
    )
    monthly_transactions = StockMovement.objects.filter(
        movement_date__date__gte=thirty_days_ago
    ).count()

    # ── Category distribution (inventory value) ───────────────────────────────
    from apps.inventory.models import Category
    cat_qs = (
        InventoryItem.objects.filter(status='Active')
        .values('category__category_name')
        .annotate(val=Sum(F('purchase_price') * F('current_qty')))
        .order_by('-val')[:8]
    )
    cat_labels = [r['category__category_name'] for r in cat_qs]
    cat_values = [float(r['val'] or 0) for r in cat_qs]
    cat_colors = [
        '#3b82f6','#10b981','#8b5cf6','#f59e0b',
        '#ef4444','#06b6d4','#84cc16','#f97316',
    ]

    # ── Asset status distribution ─────────────────────────────────────────────
    asset_qs = (
        Asset.objects.filter(is_active=True)
        .values('asset_status')
        .annotate(cnt=Count('id'))
    )
    asset_labels = [r['asset_status'] for r in asset_qs]
    asset_values = [r['cnt'] for r in asset_qs]

    # ── Stock trend last 30 days ──────────────────────────────────────────────
    chart_labels, chart_in, chart_out = [], [], []
    for i in range(29, -1, -1):
        d = today - timedelta(days=i)
        chart_labels.append(d.strftime('%b %d'))
        in_qty = (
            StockMovement.objects.filter(
                movement_date__date=d, movement_type__in=stock_in_types
            ).aggregate(s=Sum('quantity'))['s'] or 0
        )
        out_qty = (
            StockMovement.objects.filter(
                movement_date__date=d, movement_type__in=stock_out_types
            ).aggregate(s=Sum('quantity'))['s'] or 0
        )
        chart_in.append(in_qty)
        chart_out.append(out_qty)

    recent_movements = (
        StockMovement.objects.select_related('item', 'created_by')
        .order_by('-movement_date')[:10]
    )

    context = {
        'page_title': 'Reports Dashboard',
        'generated_at': timezone.now(),
        'generated_by': _get_generated_by(request),
        # KPI
        'total_items': total_items,
        'total_assets': total_assets,
        'total_inv_value': total_inv_value,
        'low_stock_count': low_stock_count,
        'stock_in_total': stock_in_total,
        'stock_out_total': stock_out_total,
        'monthly_transactions': monthly_transactions,
        # Charts
        'cat_labels': cat_labels,
        'cat_values': cat_values,
        'cat_colors': cat_colors,
        'asset_labels': asset_labels,
        'asset_values': asset_values,
        'chart_labels': chart_labels,
        'chart_in': chart_in,
        'chart_out': chart_out,
        # Recent
        'recent_movements': recent_movements,
    }
    return render(request, 'reports/index.html', context)


# Keep old name working
@login_required
@manager_or_above_required
def report_index(request):
    return report_dashboard(request)


# ──────────────────────────────────────────────────────────────────────────────
# INVENTORY REPORT
# ──────────────────────────────────────────────────────────────────────────────

def _get_inventory_queryset(request):
    from apps.inventory.models import InventoryItem
    qs = InventoryItem.objects.select_related('category', 'supplier')

    status_f  = request.GET.get('status', 'Active')
    cat_f     = request.GET.get('category', '')
    supplier_f = request.GET.get('supplier', '')
    search_f  = request.GET.get('search', '')

    if status_f:
        qs = qs.filter(status=status_f)
    if cat_f:
        qs = qs.filter(category__id=cat_f)
    if supplier_f:
        qs = qs.filter(supplier__id=supplier_f)
    if search_f:
        qs = qs.filter(
            Q(item_name__icontains=search_f) | Q(item_code__icontains=search_f)
        )
    return qs.order_by('item_name')


@login_required
@manager_or_above_required
def inventory_report(request):
    from apps.inventory.models import Category, Supplier
    items     = _get_inventory_queryset(request)
    low_stock = items.filter(current_qty__lte=F('min_qty'))
    total_value = items.aggregate(
        total=Sum(F('purchase_price') * F('current_qty'))
    )['total'] or 0
    active_suppliers = items.values('supplier').distinct().count()

    categories = Category.objects.filter(category_type='Inventory', status='Active')
    suppliers  = Supplier.objects.filter(status='Active')

    # Chart data
    cat_qs = (
        items.values('category__category_name')
        .annotate(val=Sum(F('purchase_price') * F('current_qty')))
        .order_by('-val')[:8]
    )
    cat_labels = [r['category__category_name'] for r in cat_qs]
    cat_values = [float(r['val'] or 0) for r in cat_qs]

    top10 = sorted(items, key=lambda x: x.get_total_value(), reverse=True)[:10]
    bar_labels = [i.item_name[:20] for i in top10]
    bar_values = [float(i.get_total_value()) for i in top10]

    return render(request, 'reports/inventory_report.html', {
        'items': items,
        'low_stock': low_stock,
        'total_value': total_value,
        'active_suppliers': active_suppliers,
        'categories': categories,
        'suppliers': suppliers,
        'cat_filter':   request.GET.get('category', ''),
        'sup_filter':   request.GET.get('supplier', ''),
        'status_filter': request.GET.get('status', 'Active'),
        'search_filter': request.GET.get('search', ''),
        'cat_labels': cat_labels,
        'cat_values': cat_values,
        'bar_labels': bar_labels,
        'bar_values': bar_values,
        'page_title': 'Inventory Report',
        'generated_at': timezone.now(),
        'generated_by': _get_generated_by(request),
    })


# ──────────────────────────────────────────────────────────────────────────────
# STOCK MOVEMENT REPORT
# ──────────────────────────────────────────────────────────────────────────────

def _get_stock_queryset(request):
    from apps.stock.models import StockMovement
    today        = timezone.now().date()
    default_from = today - timedelta(days=30)
    date_from    = _parse_date(request.GET.get('date_from'), default_from)
    date_to      = _parse_date(request.GET.get('date_to'), today)
    type_f       = request.GET.get('type', '')
    search_f     = request.GET.get('search', '')

    qs = StockMovement.objects.select_related('item', 'created_by').filter(
        movement_date__date__gte=date_from,
        movement_date__date__lte=date_to,
    )
    if type_f:
        qs = qs.filter(movement_type=type_f)
    if search_f:
        qs = qs.filter(
            Q(item__item_name__icontains=search_f) | Q(item__item_code__icontains=search_f)
        )
    return qs.order_by('-movement_date'), date_from, date_to


@login_required
@manager_or_above_required
def stock_movement_report(request):
    from apps.stock.models import StockMovement
    movements, date_from, date_to = _get_stock_queryset(request)
    stock_in_types  = ['Stock IN', 'Purchase']
    stock_out_types = ['Stock OUT', 'Damage', 'Lost', 'Expired', 'Usage']

    stock_in_total  = movements.filter(movement_type__in=stock_in_types).aggregate(s=Sum('quantity'))['s'] or 0
    stock_out_total = movements.filter(movement_type__in=stock_out_types).aggregate(s=Sum('quantity'))['s'] or 0
    adj_total       = movements.filter(movement_type='Adjustment').aggregate(s=Sum('quantity'))['s'] or 0

    # Daily chart
    chart_dates, chart_in_daily, chart_out_daily = [], [], []
    current = date_from
    while current <= date_to:
        chart_dates.append(current.strftime('%b %d'))
        chart_in_daily.append(
            movements.filter(movement_date__date=current, movement_type__in=stock_in_types)
            .aggregate(s=Sum('quantity'))['s'] or 0
        )
        chart_out_daily.append(
            movements.filter(movement_date__date=current, movement_type__in=stock_out_types)
            .aggregate(s=Sum('quantity'))['s'] or 0
        )
        current += timedelta(days=1)

    # Type distribution
    type_qs = movements.values('movement_type').annotate(cnt=Sum('quantity'))
    type_labels = [r['movement_type'] for r in type_qs]
    type_values = [r['cnt'] or 0 for r in type_qs]

    return render(request, 'reports/stock_report.html', {
        'movements': movements,
        'date_from': str(date_from),
        'date_to':   str(date_to),
        'type_filter': request.GET.get('type', ''),
        'search_filter': request.GET.get('search', ''),
        'movement_types': StockMovement.MOVEMENT_TYPES,
        'stock_in_total': stock_in_total,
        'stock_out_total': stock_out_total,
        'adj_total': adj_total,
        'chart_dates': chart_dates,
        'chart_in_daily': chart_in_daily,
        'chart_out_daily': chart_out_daily,
        'type_labels': type_labels,
        'type_values': type_values,
        'page_title': 'Stock Movement Report',
        'generated_at': timezone.now(),
        'generated_by': _get_generated_by(request),
    })


# ──────────────────────────────────────────────────────────────────────────────
# ASSET REPORT
# ──────────────────────────────────────────────────────────────────────────────

def _get_asset_queryset(request):
    from apps.assets.models import Asset
    qs = Asset.objects.select_related('category', 'location', 'assigned_to').filter(is_active=True)
    status_f   = request.GET.get('status', '')
    cat_f      = request.GET.get('category', '')
    location_f = request.GET.get('location', '')
    search_f   = request.GET.get('search', '')
    if status_f:
        qs = qs.filter(asset_status=status_f)
    if cat_f:
        qs = qs.filter(category__id=cat_f)
    if location_f:
        qs = qs.filter(location__id=location_f)
    if search_f:
        qs = qs.filter(
            Q(asset_name__icontains=search_f) | Q(asset_code__icontains=search_f)
        )
    return qs.order_by('asset_name')


@login_required
@manager_or_above_required
def asset_report(request):
    from apps.assets.models import Asset
    from apps.inventory.models import Category, Location
    assets = _get_asset_queryset(request)
    status_summary = Asset.objects.filter(is_active=True).values('asset_status').annotate(count=Count('id'))
    categories = Category.objects.filter(category_type='Asset', status='Active')
    locations  = Location.objects.filter(status='Active')

    available   = assets.filter(asset_status='Available').count()
    assigned    = assets.filter(asset_status='Assigned').count()
    maintenance = assets.filter(asset_status='Under Maintenance').count()

    # Chart data
    asset_status_qs = assets.values('asset_status').annotate(cnt=Count('id'))
    asset_labels = [r['asset_status'] for r in asset_status_qs]
    asset_values = [r['cnt'] for r in asset_status_qs]

    cat_qs = assets.values('category__category_name').annotate(cnt=Count('id')).order_by('-cnt')
    cat_bar_labels = [r['category__category_name'] for r in cat_qs]
    cat_bar_values = [r['cnt'] for r in cat_qs]

    return render(request, 'reports/asset_report.html', {
        'assets': assets,
        'status_filter':   request.GET.get('status', ''),
        'cat_filter':      request.GET.get('category', ''),
        'location_filter': request.GET.get('location', ''),
        'search_filter':   request.GET.get('search', ''),
        'status_choices':  Asset.STATUS_CHOICES,
        'status_summary':  status_summary,
        'categories': categories,
        'locations':  locations,
        'available':   available,
        'assigned':    assigned,
        'maintenance': maintenance,
        'asset_labels': asset_labels,
        'asset_values': asset_values,
        'cat_bar_labels': cat_bar_labels,
        'cat_bar_values': cat_bar_values,
        'page_title': 'Asset Report',
        'generated_at': timezone.now(),
        'generated_by': _get_generated_by(request),
    })


# ──────────────────────────────────────────────────────────────────────────────
# LOW STOCK REPORT
# ──────────────────────────────────────────────────────────────────────────────

def _get_low_stock_queryset(request):
    from apps.inventory.models import InventoryItem
    qs = InventoryItem.objects.select_related('category', 'supplier').filter(
        status='Active', current_qty__lte=F('min_qty')
    )
    cat_f      = request.GET.get('category', '')
    supplier_f = request.GET.get('supplier', '')
    search_f   = request.GET.get('search', '')
    if cat_f:
        qs = qs.filter(category__id=cat_f)
    if supplier_f:
        qs = qs.filter(supplier__id=supplier_f)
    if search_f:
        qs = qs.filter(
            Q(item_name__icontains=search_f) | Q(item_code__icontains=search_f)
        )
    return qs.order_by('current_qty')


@login_required
@manager_or_above_required
def low_stock_report(request):
    from apps.inventory.models import Category, Supplier
    items = _get_low_stock_queryset(request)
    out_of_stock = items.filter(current_qty=0).count()
    restock_value = sum(
        (item.min_qty - item.current_qty) * item.purchase_price for item in items
    )
    categories = Category.objects.filter(category_type='Inventory', status='Active')
    suppliers  = Supplier.objects.filter(status='Active')

    # Deficit list for template
    deficit_data = []
    for item in items:
        deficit_data.append({
            'item': item,
            'deficit': item.min_qty - item.current_qty,
        })

    # Charts
    bar_labels = [d['item'].item_name[:20] for d in deficit_data[:15]]
    bar_values = [d['item'].current_qty for d in deficit_data[:15]]

    cat_qs = items.values('category__category_name').annotate(cnt=Count('id'))
    donut_labels = [r['category__category_name'] for r in cat_qs]
    donut_values = [r['cnt'] for r in cat_qs]

    return render(request, 'reports/low_stock_report.html', {
        'items': items,
        'deficit_data': deficit_data,
        'out_of_stock': out_of_stock,
        'restock_value': restock_value,
        'categories': categories,
        'suppliers':  suppliers,
        'cat_filter':  request.GET.get('category', ''),
        'sup_filter':  request.GET.get('supplier', ''),
        'search_filter': request.GET.get('search', ''),
        'bar_labels':    bar_labels,
        'bar_values':    bar_values,
        'donut_labels':  donut_labels,
        'donut_values':  donut_values,
        'page_title': 'Low Stock Report',
        'generated_at': timezone.now(),
        'generated_by': _get_generated_by(request),
    })


# ──────────────────────────────────────────────────────────────────────────────
# EXPORT HELPERS
# ──────────────────────────────────────────────────────────────────────────────

def _excel_styled_wb(title):
    """Return (wb, ws) with a styled header row ready."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title
    ws.header_style_font  = Font(name='Calibri', bold=True, color='FFFFFF')
    ws.header_style_fill  = PatternFill('solid', fgColor='2563EB')
    ws.header_style_align = Alignment(horizontal='center', vertical='center')
    return wb, ws


def _apply_header_row(ws, headers):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    font  = Font(name='Calibri', bold=True, color='FFFFFF')
    fill  = PatternFill('solid', fgColor='2563EB')
    align = Alignment(horizontal='center', vertical='center')
    ws.append(headers)
    for cell in ws[1]:
        cell.font      = font
        cell.fill      = fill
        cell.alignment = align


def _auto_col_widths(ws):
    for col in ws.columns:
        max_len = max((len(str(c.value or '')) for c in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)


def _excel_response(wb, filename):
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


def _csv_response(filename):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def _pdf_canvas(buffer, title, generated_by, generated_at, total_records):
    """Return a ReportLab canvas with header already drawn. Caller draws body then saves."""
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.pdfgen import canvas as rl_canvas
    from reportlab.lib import colors
    from reportlab.lib.units import cm

    page_w, page_h = landscape(A4)
    c = rl_canvas.Canvas(buffer, pagesize=landscape(A4))
    c._eiams_title    = title
    c._eiams_by       = generated_by
    c._eiams_at       = generated_at.strftime('%Y-%m-%d %H:%M')
    c._eiams_total    = total_records
    c._eiams_page_w   = page_w
    c._eiams_page_h   = page_h
    return c


def _draw_pdf_header(c, title):
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    w = c._eiams_page_w
    h = c._eiams_page_h
    # Blue banner
    c.setFillColor(colors.HexColor('#2563EB'))
    c.rect(0, h - 2.2*cm, w, 2.2*cm, fill=1, stroke=0)
    c.setFillColor(colors.white)
    c.setFont('Helvetica-Bold', 14)
    c.drawString(1*cm, h - 1.2*cm, 'EIAMS — Enterprise Inventory & Asset Management System')
    c.setFont('Helvetica', 10)
    c.drawString(1*cm, h - 1.8*cm, f'{title}   |   Generated: {c._eiams_at}   |   By: {c._eiams_by}   |   Records: {c._eiams_total}')


def _draw_pdf_footer(c, page_num):
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    w = c._eiams_page_w
    c.setFillColor(colors.HexColor('#64748b'))
    c.setFont('Helvetica', 8)
    c.drawString(1*cm, 0.6*cm, f'EIAMS — Confidential   |   Page {page_num}')
    c.drawRightString(w - 1*cm, 0.6*cm, 'Total: {} records'.format(c._eiams_total))


# ──────────────────────────────────────────────────────────────────────────────
# INVENTORY EXPORTS
# ──────────────────────────────────────────────────────────────────────────────

@login_required
@manager_or_above_required
def export_inventory_excel(request):
    items = _get_inventory_queryset(request)
    wb, ws = _excel_styled_wb('Inventory')
    _apply_header_row(ws, ['#', 'Item Code', 'Item Name', 'Category', 'Supplier', 'Unit',
                            'Min Qty', 'Current Qty', 'Unit Price', 'Total Value', 'Status'])
    for i, item in enumerate(items, 1):
        ws.append([
            i, item.item_code, item.item_name,
            item.category.category_name,
            item.supplier.supplier_name if item.supplier else '',
            item.unit, item.min_qty, item.current_qty,
            float(item.purchase_price), float(item.get_total_value()), item.status,
        ])
    _auto_col_widths(ws)
    return _excel_response(wb, 'inventory_report.xlsx')


@login_required
@manager_or_above_required
def export_inventory_csv(request):
    response = _csv_response('inventory_report.csv')
    writer = csv.writer(response)
    writer.writerow(['#', 'Item Code', 'Item Name', 'Category', 'Supplier', 'Unit',
                     'Min Qty', 'Current Qty', 'Unit Price', 'Total Value', 'Status'])
    for i, item in enumerate(_get_inventory_queryset(request), 1):
        writer.writerow([
            i, item.item_code, item.item_name,
            item.category.category_name,
            item.supplier.supplier_name if item.supplier else '',
            item.unit, item.min_qty, item.current_qty,
            item.purchase_price, item.get_total_value(), item.status,
        ])
    return response


@login_required
@manager_or_above_required
def export_inventory_pdf(request):
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.pagesizes import landscape, A4

    items = list(_get_inventory_queryset(request))
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                            leftMargin=1*cm, rightMargin=1*cm,
                            topMargin=2.5*cm, bottomMargin=1.5*cm)
    styles = getSampleStyleSheet()
    generated_by = _get_generated_by(request)
    generated_at = timezone.now()

    header_data = [['#', 'Item Code', 'Item Name', 'Category', 'Supplier',
                     'Unit', 'Min', 'Qty', 'Price', 'Value', 'Status']]
    for i, item in enumerate(items, 1):
        header_data.append([
            i, item.item_code, item.item_name[:22],
            item.category.category_name[:15],
            (item.supplier.supplier_name[:15] if item.supplier else '—'),
            item.unit, item.min_qty, item.current_qty,
            f'${item.purchase_price:.2f}',
            f'${item.get_total_value():.2f}',
            item.status,
        ])

    t = Table(header_data, repeatRows=1)
    t.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#2563EB')),
        ('TEXTCOLOR',  (0,0), (-1,0), colors.white),
        ('FONTNAME',   (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE',   (0,0), (-1,-1), 8),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#f1f5f9')]),
        ('GRID', (0,0), (-1,-1), 0.3, colors.HexColor('#cbd5e1')),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
    ]))

    title_p = Paragraph(f'<b>Inventory Stock Report</b> — EIAMS', styles['Title'])
    sub_p   = Paragraph(f'Generated: {generated_at.strftime("%Y-%m-%d %H:%M")} | By: {generated_by} | Total: {len(items)} records', styles['Normal'])

    doc.build([title_p, Spacer(1, 0.3*cm), sub_p, Spacer(1, 0.5*cm), t])
    buf.seek(0)
    response = HttpResponse(buf, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="inventory_report.pdf"'
    return response


# ──────────────────────────────────────────────────────────────────────────────
# STOCK MOVEMENT EXPORTS
# ──────────────────────────────────────────────────────────────────────────────

@login_required
@manager_or_above_required
def export_stock_excel(request):
    movements, _, _ = _get_stock_queryset(request)
    wb, ws = _excel_styled_wb('Stock Movements')
    _apply_header_row(ws, ['#', 'Date', 'Item Code', 'Item Name', 'Movement Type',
                            'Qty', 'Qty After', 'Reference', 'Recorded By', 'Remarks'])
    for i, m in enumerate(movements, 1):
        ws.append([
            i,
            m.movement_date.strftime('%Y-%m-%d %H:%M'),
            m.item.item_code, m.item.item_name,
            m.movement_type, m.quantity,
            m.qty_after if m.qty_after is not None else '',
            m.reference_no or '',
            m.created_by.full_name,
            m.remarks or '',
        ])
    _auto_col_widths(ws)
    return _excel_response(wb, 'stock_movement_report.xlsx')


@login_required
@manager_or_above_required
def export_stock_csv(request):
    movements, _, _ = _get_stock_queryset(request)
    response = _csv_response('stock_movement_report.csv')
    writer = csv.writer(response)
    writer.writerow(['#', 'Date', 'Item Code', 'Item Name', 'Movement Type',
                     'Qty', 'Qty After', 'Reference', 'Recorded By', 'Remarks'])
    for i, m in enumerate(movements, 1):
        writer.writerow([
            i, m.movement_date.strftime('%Y-%m-%d %H:%M'),
            m.item.item_code, m.item.item_name,
            m.movement_type, m.quantity,
            m.qty_after if m.qty_after is not None else '',
            m.reference_no or '',
            m.created_by.full_name,
            m.remarks or '',
        ])
    return response


@login_required
@manager_or_above_required
def export_stock_pdf(request):
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.pagesizes import landscape, A4

    movements, date_from, date_to = _get_stock_queryset(request)
    movements = list(movements)
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                            leftMargin=1*cm, rightMargin=1*cm,
                            topMargin=2.5*cm, bottomMargin=1.5*cm)
    styles = getSampleStyleSheet()
    generated_by = _get_generated_by(request)
    generated_at = timezone.now()

    data = [['#', 'Date', 'Item Code', 'Item Name', 'Type', 'Qty', 'Qty After', 'Reference', 'Recorded By']]
    for i, m in enumerate(movements, 1):
        data.append([
            i,
            m.movement_date.strftime('%Y-%m-%d'),
            m.item.item_code,
            m.item.item_name[:20],
            m.movement_type,
            m.quantity,
            m.qty_after if m.qty_after is not None else '—',
            m.reference_no or '—',
            m.created_by.full_name[:15],
        ])

    t = Table(data, repeatRows=1)
    t.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#059669')),
        ('TEXTCOLOR',  (0,0), (-1,0), colors.white),
        ('FONTNAME',   (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE',   (0,0), (-1,-1), 8),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#f0fdf4')]),
        ('GRID', (0,0), (-1,-1), 0.3, colors.HexColor('#cbd5e1')),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
    ]))

    title_p = Paragraph(f'<b>Stock Movement Report</b> — EIAMS ({date_from} to {date_to})', styles['Title'])
    sub_p   = Paragraph(f'Generated: {generated_at.strftime("%Y-%m-%d %H:%M")} | By: {generated_by} | Total: {len(movements)} records', styles['Normal'])
    doc.build([title_p, Spacer(1, 0.3*cm), sub_p, Spacer(1, 0.5*cm), t])
    buf.seek(0)
    response = HttpResponse(buf, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="stock_movement_report.pdf"'
    return response


# ──────────────────────────────────────────────────────────────────────────────
# ASSET EXPORTS
# ──────────────────────────────────────────────────────────────────────────────

@login_required
@manager_or_above_required
def export_asset_excel(request):
    assets = _get_asset_queryset(request)
    wb, ws = _excel_styled_wb('Assets')
    _apply_header_row(ws, ['#', 'Asset Code', 'Asset Name', 'Category', 'Location',
                            'Assigned To', 'Purchase Date', 'Purchase Price',
                            'Warranty Expiry', 'Status'])
    for i, a in enumerate(assets, 1):
        ws.append([
            i, a.asset_code, a.asset_name,
            a.category.category_name,
            a.location.location_name if a.location else '',
            a.assigned_to.full_name if a.assigned_to else '',
            str(a.purchase_date) if a.purchase_date else '',
            float(a.purchase_price),
            str(a.warranty_expiry_date) if a.warranty_expiry_date else '',
            a.asset_status,
        ])
    _auto_col_widths(ws)
    return _excel_response(wb, 'asset_report.xlsx')


@login_required
@manager_or_above_required
def export_asset_csv(request):
    response = _csv_response('asset_report.csv')
    writer = csv.writer(response)
    writer.writerow(['#', 'Asset Code', 'Asset Name', 'Category', 'Location',
                     'Assigned To', 'Purchase Date', 'Purchase Price', 'Warranty Expiry', 'Status'])
    for i, a in enumerate(_get_asset_queryset(request), 1):
        writer.writerow([
            i, a.asset_code, a.asset_name,
            a.category.category_name,
            a.location.location_name if a.location else '',
            a.assigned_to.full_name if a.assigned_to else '',
            str(a.purchase_date) if a.purchase_date else '',
            a.purchase_price,
            str(a.warranty_expiry_date) if a.warranty_expiry_date else '',
            a.asset_status,
        ])
    return response


@login_required
@manager_or_above_required
def export_asset_pdf(request):
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.pagesizes import landscape, A4

    assets = list(_get_asset_queryset(request))
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                            leftMargin=1*cm, rightMargin=1*cm,
                            topMargin=2.5*cm, bottomMargin=1.5*cm)
    styles = getSampleStyleSheet()
    generated_by = _get_generated_by(request)
    generated_at = timezone.now()

    data = [['#', 'Asset Code', 'Asset Name', 'Category', 'Location',
             'Assigned To', 'Purchase Date', 'Price', 'Status']]
    for i, a in enumerate(assets, 1):
        data.append([
            i, a.asset_code, a.asset_name[:22],
            a.category.category_name[:15],
            (a.location.location_name[:15] if a.location else '—'),
            (a.assigned_to.full_name[:15] if a.assigned_to else '—'),
            str(a.purchase_date) if a.purchase_date else '—',
            f'${a.purchase_price:.2f}',
            a.asset_status,
        ])

    t = Table(data, repeatRows=1)
    t.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#7c3aed')),
        ('TEXTCOLOR',  (0,0), (-1,0), colors.white),
        ('FONTNAME',   (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE',   (0,0), (-1,-1), 8),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#f5f3ff')]),
        ('GRID', (0,0), (-1,-1), 0.3, colors.HexColor('#cbd5e1')),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
    ]))

    title_p = Paragraph('<b>Asset Report</b> — EIAMS', styles['Title'])
    sub_p   = Paragraph(f'Generated: {generated_at.strftime("%Y-%m-%d %H:%M")} | By: {generated_by} | Total: {len(assets)} records', styles['Normal'])
    doc.build([title_p, Spacer(1, 0.3*cm), sub_p, Spacer(1, 0.5*cm), t])
    buf.seek(0)
    response = HttpResponse(buf, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="asset_report.pdf"'
    return response


# ──────────────────────────────────────────────────────────────────────────────
# LOW STOCK EXPORTS
# ──────────────────────────────────────────────────────────────────────────────

@login_required
@manager_or_above_required
def export_low_stock_excel(request):
    items = _get_low_stock_queryset(request)
    wb, ws = _excel_styled_wb('Low Stock')
    _apply_header_row(ws, ['#', 'Item Code', 'Item Name', 'Category', 'Supplier',
                            'Unit', 'Min Qty', 'Current Qty', 'Deficit', 'Unit Price', 'Urgency'])
    for i, item in enumerate(items, 1):
        deficit = item.min_qty - item.current_qty
        urgency = 'Out of Stock' if item.current_qty == 0 else ('Critical' if deficit > item.min_qty / 2 else 'Low')
        ws.append([
            i, item.item_code, item.item_name,
            item.category.category_name,
            item.supplier.supplier_name if item.supplier else '',
            item.unit, item.min_qty, item.current_qty,
            deficit, float(item.purchase_price), urgency,
        ])
    _auto_col_widths(ws)
    return _excel_response(wb, 'low_stock_report.xlsx')


@login_required
@manager_or_above_required
def export_low_stock_csv(request):
    response = _csv_response('low_stock_report.csv')
    writer = csv.writer(response)
    writer.writerow(['#', 'Item Code', 'Item Name', 'Category', 'Supplier',
                     'Unit', 'Min Qty', 'Current Qty', 'Deficit', 'Unit Price', 'Urgency'])
    for i, item in enumerate(_get_low_stock_queryset(request), 1):
        deficit = item.min_qty - item.current_qty
        urgency = 'Out of Stock' if item.current_qty == 0 else ('Critical' if deficit > item.min_qty / 2 else 'Low')
        writer.writerow([
            i, item.item_code, item.item_name,
            item.category.category_name,
            item.supplier.supplier_name if item.supplier else '',
            item.unit, item.min_qty, item.current_qty,
            deficit, item.purchase_price, urgency,
        ])
    return response


@login_required
@manager_or_above_required
def export_low_stock_pdf(request):
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.pagesizes import landscape, A4

    items = list(_get_low_stock_queryset(request))
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                            leftMargin=1*cm, rightMargin=1*cm,
                            topMargin=2.5*cm, bottomMargin=1.5*cm)
    styles = getSampleStyleSheet()
    generated_by = _get_generated_by(request)
    generated_at = timezone.now()

    data = [['#', 'Item Code', 'Item Name', 'Category', 'Supplier',
             'Unit', 'Min', 'Current', 'Deficit', 'Unit Price', 'Urgency']]
    for i, item in enumerate(items, 1):
        deficit = item.min_qty - item.current_qty
        urgency = 'Out of Stock' if item.current_qty == 0 else ('Critical' if deficit > item.min_qty / 2 else 'Low')
        data.append([
            i, item.item_code, item.item_name[:22],
            item.category.category_name[:15],
            (item.supplier.supplier_name[:15] if item.supplier else '—'),
            item.unit, item.min_qty, item.current_qty,
            deficit, f'${item.purchase_price:.2f}', urgency,
        ])

    t = Table(data, repeatRows=1)
    t.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#dc2626')),
        ('TEXTCOLOR',  (0,0), (-1,0), colors.white),
        ('FONTNAME',   (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE',   (0,0), (-1,-1), 8),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#fef2f2')]),
        ('GRID', (0,0), (-1,-1), 0.3, colors.HexColor('#cbd5e1')),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
    ]))

    title_p = Paragraph('<b>Low Stock Report</b> — EIAMS', styles['Title'])
    sub_p   = Paragraph(f'Generated: {generated_at.strftime("%Y-%m-%d %H:%M")} | By: {generated_by} | Total: {len(items)} records', styles['Normal'])
    doc.build([title_p, Spacer(1, 0.3*cm), sub_p, Spacer(1, 0.5*cm), t])
    buf.seek(0)
    response = HttpResponse(buf, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="low_stock_report.pdf"'
    return response
